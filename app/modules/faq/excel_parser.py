"""
Utility to parse Excel files into FAQ rows for bulk import.
Uses openpyxl for in-memory processing.
"""
import io
import openpyxl
import logging
from typing import List, Dict, Any, Optional, Tuple, Union
import re
from openpyxl.cell.rich_text import CellRichText
from app.core.format_utils import rich_text_to_markdown
from app.integrations.excel import get_column_index, cell_to_rich_text, get_cell_value
from app.modules.metadata.models import YEAR_MIN, YEAR_MAX
from app.modules.metadata.utils.parsers import parse_year_range

logger = logging.getLogger(__name__)


def parse_excel_to_faq_rows(
    file_bytes: bytes,
    question_col: str,
    answer_col: str,
    metadata_map: Optional[Dict[str, str]] = None,
    sheet_name: Optional[str] = None,
    skip_rows: int = 1, # Default skip header
) -> Dict[str, Any]:
    """
    Parse Excel bytes and return a list of rows with question, answer, and metadata.
    Formatting (bold, italic, underline, hyperlinks) is preserved as HTML Rich Text.
    """
    try:
        # data_only=False + rich_text=True is the key to preserving formatting
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=False, rich_text=True)
        if sheet_name:
            if sheet_name not in wb.sheetnames:
                raise ValueError(f"Sheet '{sheet_name}' not found in workbook.")
            sheet = wb[sheet_name]
        else:
            sheet = wb.active
            
        q_idx = get_column_index(sheet, question_col, header_row=skip_rows)
        a_idx = get_column_index(sheet, answer_col, header_row=skip_rows)
        
        meta_indices = {}
        if metadata_map:
            for key, col in metadata_map.items():
                idx = get_column_index(sheet, col, header_row=skip_rows)
                if idx:
                    meta_indices[key] = idx
        
        if not q_idx:
            raise ValueError(f"Could not find question column: '{question_col}'")
        if not a_idx:
            raise ValueError(f"Could not find answer column: '{answer_col}'")
            
        rows = []
        valid_count = 0
        invalid_count = 0
        
        # Iterate starting from skip_rows + 1
        for i, row in enumerate(sheet.iter_rows(min_row=skip_rows + 1), start=skip_rows + 1):
            # Use cell_to_rich_text to preserve formatting as HTML
            q_val = cell_to_rich_text(row[q_idx - 1])
            a_val = cell_to_rich_text(row[a_idx - 1])
            
            # Extract metadata
            # Initialize metadata with default ranges using constants
            metadata = {
                "enrollment_year": {"from_year": YEAR_MIN, "to_year": YEAR_MAX},
                "academic_year": {"from_year": YEAR_MIN, "to_year": YEAR_MAX}
            }
            
            for key, idx in meta_indices.items():
                clean_val = get_cell_value(row[idx - 1]) if idx <= len(row) else None
                
                # Handle both snake_case and camelCase from mapping
                if key in ["enrollment_year", "enrollmentYear", "academic_year", "academicYear"]:
                    if not clean_val:
                        continue
                    rng = parse_year_range(str(clean_val))
                    out_key = "enrollment_year" if key in ["enrollment_year", "enrollmentYear"] else "academic_year"
                    metadata[out_key] = {"from_year": rng.from_year, "to_year": rng.to_year}
            
            # Basic validation
            error = None
            is_valid = True
            
            # Strip HTML tags for length validation
            clean_q = re.sub(r'<[^>]+>', '', q_val).strip()
            
            if not q_val or not clean_q:
                error = "Question is missing"
                is_valid = False
            elif not a_val or not re.sub(r'<[^>]+>', '', a_val).strip():
                error = "Answer is missing"
                is_valid = False
            elif len(clean_q) < 5:
                error = "Question too short (min 5 chars)"
                is_valid = False
                
            row_data = {
                "row_index": i,
                "question": q_val,
                "answer_rich_text": a_val,
                "answer_markdown": rich_text_to_markdown(a_val),
                "metadata": metadata,
                "is_valid": is_valid,
                "error": error
            }
            
            rows.append(row_data)
            if is_valid:
                valid_count += 1
            else:
                invalid_count += 1
                
        return {
            "rows": rows,
            "total_rows": len(rows),
            "valid_rows": valid_count,
            "invalid_rows": invalid_count
        }
        
    except Exception as e:
        logger.error(f"Error parsing Excel: {e}")
        raise ValueError(f"Failed to parse Excel file: {str(e)}")
